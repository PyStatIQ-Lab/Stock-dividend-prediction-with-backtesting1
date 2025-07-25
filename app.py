import yfinance as yf
import pandas as pd
import os
import numpy as np
from openpyxl import load_workbook
import streamlit as st
import matplotlib.pyplot as plt
import seaborn as sns
from datetime import datetime, timedelta

# Configure matplotlib
plt.rcParams['font.family'] = 'DejaVu Sans'
plt.rcParams['axes.facecolor'] = '#0E1117'
plt.rcParams['axes.edgecolor'] = 'white'
plt.rcParams['axes.labelcolor'] = 'white'
plt.rcParams['xtick.color'] = 'white'
plt.rcParams['ytick.color'] = 'white'
plt.rcParams['text.color'] = 'white'

# Path to the stocks.xlsx file
STOCKS_FILE_PATH = 'stocks.xlsx'  # Change this to the correct path if needed

# Function to fetch data for a given stock ticker
def get_financial_data(ticker):
    stock = yf.Ticker(ticker)
    result = {'Ticker': ticker}
    
    try:
        income_statement = stock.financials
        balance_sheet = stock.balance_sheet
        cash_flow = stock.cashflow
        dividends = stock.dividends
    except Exception as e:
        st.error(f"Error fetching financial data for {ticker}: {e}")
        return None

    try:
        historical_data = stock.history(period="1d")
        latest_close_price = historical_data['Close'].iloc[-1]
    except Exception as e:
        latest_close_price = "N/A"

    result['Net Income'] = income_statement.loc['Net Income'] if 'Net Income' in income_statement.index else "N/A"
    result['Operating Income'] = income_statement.loc['Operating Income'] if 'Operating Income' in income_statement.index else \
                                 income_statement.loc['EBIT'] if 'EBIT' in income_statement.index else "N/A"
    
    try:
        shares = stock.info.get('sharesOutstanding', None)
        if shares and 'Earnings Before Interest and Taxes' in income_statement.index:
            eps = income_statement.loc['Earnings Before Interest and Taxes'] / shares
        else:
            eps = "N/A"
    except Exception:
        eps = "N/A"
    result['EPS'] = eps
    
    if 'Total Revenue' in income_statement.index:
        revenue = income_statement.loc['Total Revenue']
        if len(revenue) > 1:
            result['Revenue Growth'] = revenue.pct_change().iloc[-1] * 100
        else:
            result['Revenue Growth'] = "N/A"
    else:
        result['Revenue Growth'] = "N/A"
    
    result['Retained Earnings'] = balance_sheet.loc['Retained Earnings'] if 'Retained Earnings' in balance_sheet.index else "N/A"
    result['Cash Reserves'] = balance_sheet.loc['Cash'] if 'Cash' in balance_sheet.index else "N/A"
    
    try:
        total_debt = balance_sheet.loc['Total Debt'] if 'Total Debt' in balance_sheet.index else 0
        stockholders_equity = balance_sheet.loc['Stockholders Equity'] if 'Stockholders Equity' in balance_sheet.index else 1
        result['Debt-to-Equity Ratio'] = total_debt / stockholders_equity
    except KeyError:
        result['Debt-to-Equity Ratio'] = "N/A"
    
    try:
        total_assets = balance_sheet.loc['Total Assets']
        total_liabilities = balance_sheet.loc['Total Liabilities Net Minority Interest']
        result['Working Capital'] = total_assets - total_liabilities
    except KeyError:
        result['Working Capital'] = "N/A"
    
    result['Dividend Yield'] = stock.info.get('dividendYield', "N/A")
    
    result['Free Cash Flow'] = cash_flow.loc['Free Cash Flow'] if 'Free Cash Flow' in cash_flow.index else "N/A"
    
    if not dividends.empty:
        result['Dividend Growth Rate'] = dividends.pct_change().mean() * 100
    else:
        result['Dividend Growth Rate'] = "N/A"
    
    result['Latest Close Price'] = latest_close_price
    result['Dividend Percentage'] = "N/A"
    
    if not dividends.empty:
        predicted_dividend_amount = dividends.iloc[-1]
        if latest_close_price != "N/A":
            dividend_percentage = (predicted_dividend_amount / latest_close_price) * 100
            result['Dividend Percentage'] = dividend_percentage
        
        past_dividends = dividends.tail(10)
        result['Past Dividends'] = past_dividends.tolist()
        
        date_diffs = past_dividends.index.to_series().diff().dropna()
        if not date_diffs.empty:
            avg_diff = date_diffs.mean()
            last_dividend_date = past_dividends.index[-1]
            next_dividend_date = last_dividend_date + avg_diff
            result['Next Dividend Date'] = str(next_dividend_date.date())
        else:
            result['Next Dividend Date'] = 'N/A'

        result['Predicted Dividend Amount'] = predicted_dividend_amount
    else:
        result['Next Dividend Date'] = 'N/A'
        result['Predicted Dividend Amount'] = 'N/A'
        result['Dividend Percentage'] = "N/A"

    return result

# Backtesting function
def backtest_dividend_predictions(ticker):
    stock = yf.Ticker(ticker)
    dividends = stock.dividends.sort_index(ascending=True)
    
    if len(dividends) < 11:
        return None, f"Not enough dividend history for {ticker}. Minimum 11 dividend payments required for backtesting."
    
    results = []
    # Start from 10th dividend to have enough history
    for i in range(10, len(dividends)):
        historical = dividends.iloc[:i]
        
        # Calculate average dividend interval
        date_diffs = historical.index.to_series().diff().dropna()
        avg_interval = date_diffs.mean()
        
        # Predict next dividend
        last_date = historical.index[-1]
        predicted_date = last_date + avg_interval
        predicted_amount = historical.iloc[-1]
        
        # Get actual next dividend (if exists)
        if i < len(dividends) - 1:
            actual_date = dividends.index[i+1]
            actual_amount = dividends.iloc[i+1]
            date_error = (actual_date - predicted_date).days
            amount_error = actual_amount - predicted_amount
            pct_amount_error = (amount_error / actual_amount) * 100
        else:
            actual_date, actual_amount, date_error, amount_error, pct_amount_error = None, None, None, None, None
        
        results.append({
            'As of Date': last_date.strftime('%Y-%m-%d'),
            'Predicted Date': predicted_date.strftime('%Y-%m-%d'),
            'Predicted Amount': predicted_amount,
            'Actual Date': actual_date.strftime('%Y-%m-%d') if actual_date else "N/A",
            'Actual Amount': actual_amount if actual_amount else "N/A",
            'Date Error (days)': date_error if date_error is not None else "N/A",
            'Amount Error': amount_error if amount_error is not None else "N/A",
            'Amount Error (%)': pct_amount_error if pct_amount_error is not None else "N/A"
        })
    
    return pd.DataFrame(results), None

# Function to save results to an Excel file
def save_to_excel(results, filename="dividend_predictions.xlsx"):
    try:
        results_df = pd.DataFrame(results)
        if os.path.exists(filename):
            book = load_workbook(filename)
            writer = pd.ExcelWriter(filename, engine='openpyxl')
            writer.book = book
            results_df.to_excel(writer, index=False, header=False, startrow=book.active.max_row)
            writer.save()
        else:
            results_df.to_excel(filename, index=False)
        st.success(f"Results saved to {filename}")
    except Exception as e:
        st.error(f"Error saving to Excel: {e}")

# Streamlit App
st.set_page_config(page_title="Stock Dividend Predictions", layout="wide", page_icon="📈")

# Display Header Logo
st.markdown("""
    <style>
        .header-logo {
            display: block;
            margin-left: auto;
            margin-right: auto;
            width: 25%;
        }
        /* Hide GitHub icons and fork button */
        .css-1v0mbdj { 
            display: none !important;
        }
        .css-1b22hs3 {
            display: none !important;
        }
        /* Hide Streamlit footer elements */
        footer { 
            display: none !important; 
        }
        /* Hide the GitHub repository button */
        .css-1r6ntm8 { 
            display: none !important;
        }
        .metric-box {
            background-color: #1a1a1a;
            border-radius: 10px;
            padding: 15px;
            margin: 10px 0;
            box-shadow: 0 4px 8px rgba(0,0,0,0.1);
        }
        .metric-title {
            font-size: 16px;
            font-weight: bold;
            color: #9e9e9e;
        }
        .metric-value {
            font-size: 24px;
            font-weight: bold;
            color: #4CAF50;
        }
        .negative-value {
            color: #F44336;
        }
        .stProgress > div > div > div > div {
            background-color: #4CAF50;
        }
        .stDataFrame {
            border-radius: 10px;
        }
        .stButton>button {
            background-color: #4CAF50;
            color: white;
            border-radius: 5px;
            border: none;
            padding: 10px 24px;
            font-weight: bold;
        }
        .stButton>button:hover {
            background-color: #45a049;
        }
        .stSelectbox>div>div>div>div>div {
            background-color: #1a1a1a;
            color: white;
        }
        .stMarkdown h1 {
            color: #4CAF50;
        }
        .stMarkdown h2 {
            color: #4CAF50;
            border-bottom: 2px solid #4CAF50;
            padding-bottom: 5px;
        }
        .stMarkdown h3 {
            color: #4CAF50;
        }
    </style>
    <img class="header-logo" src="https://pystatiq.com/images/pystatIQ_logo.png" alt="Header Logo">
""", unsafe_allow_html=True)

st.title('📈 Stock Dividend Prediction & Backtesting')

# Read the stock symbols from the local stocks.xlsx file
if os.path.exists(STOCKS_FILE_PATH):
    symbols_df = pd.read_excel(STOCKS_FILE_PATH)

    # Check if the 'Symbol' column exists
    if 'Symbol' not in symbols_df.columns:
        st.error("The file must contain a 'Symbol' column with stock tickers.")
    else:
        # Let the user select stocks from the file
        stock_options = symbols_df['Symbol'].tolist()
        selected_stocks = st.multiselect("Select Stock Symbols", stock_options, help="Choose one or more stocks to analyze")

        # Button to start the data fetching process
        if st.button('📊 Fetch Financial Data', key='fetch_data') and selected_stocks:
            all_results = []
            progress_bar = st.progress(0)
            status_text = st.empty()
            
            for idx, ticker in enumerate(selected_stocks):
                progress = (idx + 1) / len(selected_stocks)
                status_text.text(f"Processing {ticker} ({idx+1}/{len(selected_stocks)})...")
                progress_bar.progress(progress)
                
                result = get_financial_data(ticker)
                if result is not None:
                    all_results.append(result)
            
            progress_bar.empty()
            status_text.empty()
            
            if all_results:
                st.subheader("📋 Financial Analysis Results")
                results_df = pd.DataFrame(all_results)
                
                # Select only the most important columns for display
                display_cols = ['Ticker', 'Latest Close Price', 'Dividend Yield', 
                               'Dividend Percentage', 'Next Dividend Date', 
                               'Predicted Dividend Amount', 'Dividend Growth Rate',
                               'Revenue Growth', 'Debt-to-Equity Ratio']
                
                # Filter columns that actually exist in the results
                available_cols = [col for col in display_cols if col in results_df.columns]
                st.dataframe(results_df[available_cols])
                
                # Button to save the results to Excel
                if st.button('💾 Save Results to Excel', key='save_results'):
                    save_to_excel(all_results)

else:
    st.error(f"{STOCKS_FILE_PATH} not found. Please ensure the file exists.")

# Backtesting Section
st.header("🔍 Backtesting Dividend Predictions")

if 'selected_stocks' in locals() and selected_stocks:
    backtest_ticker = st.selectbox("Select stock for backtesting", selected_stocks, key='backtest_select')
    
    if st.button("🔬 Run Backtest", key='run_backtest'):
        with st.spinner(f"Backtesting {backtest_ticker}..."):
            backtest_df, error = backtest_dividend_predictions(backtest_ticker)
            
            if error:
                st.error(error)
            else:
                st.success("Backtesting complete!")
                
                # Display metrics
                if not backtest_df.empty:
                    date_errors = pd.to_numeric(backtest_df['Date Error (days)'], errors='coerce').dropna()
                    amount_errors = pd.to_numeric(backtest_df['Amount Error (%)'], errors='coerce').dropna()
                    
                    col1, col2, col3 = st.columns(3)
                    
                    with col1:
                        st.markdown("<div class='metric-box'>", unsafe_allow_html=True)
                        st.markdown("<div class='metric-title'>Avg. Date Error (days)</div>", unsafe_allow_html=True)
                        avg_date_error = date_errors.mean()
                        error_class = "negative-value" if avg_date_error > 0 else ""
                        st.markdown(f"<div class='metric-value {error_class}'>{avg_date_error:.1f}</div>", unsafe_allow_html=True)
                        st.markdown("</div>", unsafe_allow_html=True)
                    
                    with col2:
                        st.markdown("<div class='metric-box'>", unsafe_allow_html=True)
                        st.markdown("<div class='metric-title'>Avg. Amount Error (%)</div>", unsafe_allow_html=True)
                        avg_amount_error = amount_errors.mean()
                        error_class = "negative-value" if avg_amount_error < 0 else ""
                        st.markdown(f"<div class='metric-value {error_class}'>{avg_amount_error:.2f}%</div>", unsafe_allow_html=True)
                        st.markdown("</div>", unsafe_allow_html=True)
                    
                    with col3:
                        st.markdown("<div class='metric-box'>", unsafe_allow_html=True)
                        st.markdown("<div class='metric-title'>Prediction Accuracy</div>", unsafe_allow_html=True)
                        accuracy = 100 - abs(avg_amount_error)
                        st.markdown(f"<div class='metric-value'>{accuracy:.1f}%</div>", unsafe_allow_html=True)
                        st.markdown("</div>", unsafe_allow_html=True)
                
                # Show results table
                st.subheader("Backtest Results")
                st.dataframe(backtest_df)
                
                # Visualizations
                st.subheader("📈 Error Analysis")
                
                if not backtest_df.empty:
                    # Convert dates to datetime for plotting
                    plot_df = backtest_df.copy()
                    plot_df['As of Date'] = pd.to_datetime(plot_df['As of Date'])
                    
                    # Create tabs for different visualizations
                    tab1, tab2, tab3 = st.tabs(["Date Errors", "Amount Errors", "Prediction Accuracy"])
                    
                    with tab1:
                        fig, ax = plt.subplots(figsize=(10, 6))
                        sns.lineplot(data=plot_df, x='As of Date', y='Date Error (days)', ax=ax, marker='o')
                        ax.axhline(y=0, color='r', linestyle='--')
                        ax.set_title('Date Prediction Error Over Time', fontsize=16)
                        ax.set_xlabel('Simulation Date')
                        ax.set_ylabel('Error (days)')
                        ax.grid(True, linestyle='--', alpha=0.7)
                        st.pyplot(fig)
                    
                    with tab2:
                        fig, ax = plt.subplots(figsize=(10, 6))
                        sns.lineplot(data=plot_df, x='As of Date', y='Amount Error (%)', ax=ax, marker='o')
                        ax.axhline(y=0, color='r', linestyle='--')
                        ax.set_title('Amount Prediction Error Over Time', fontsize=16)
                        ax.set_xlabel('Simulation Date')
                        ax.set_ylabel('Error (%)')
                        ax.grid(True, linestyle='--', alpha=0.7)
                        st.pyplot(fig)
                    
                    with tab3:
                        # Calculate prediction accuracy
                        plot_df['Date Accuracy'] = 100 - np.abs(plot_df['Date Error (days)'] / 30) * 100
                        plot_df['Amount Accuracy'] = 100 - np.abs(plot_df['Amount Error (%)'])
                        
                        fig, ax = plt.subplots(figsize=(10, 6))
                        sns.lineplot(data=plot_df, x='As of Date', y='Date Accuracy', ax=ax, label='Date Accuracy', marker='o')
                        sns.lineplot(data=plot_df, x='As of Date', y='Amount Accuracy', ax=ax, label='Amount Accuracy', marker='o')
                        ax.set_title('Prediction Accuracy Over Time', fontsize=16)
                        ax.set_xlabel('Simulation Date')
                        ax.set_ylabel('Accuracy (%)')
                        ax.legend()
                        ax.grid(True, linestyle='--', alpha=0.7)
                        ax.set_ylim(0, 100)
                        st.pyplot(fig)
                
                # Download button
                st.subheader("📥 Download Results")
                csv = backtest_df.to_csv(index=False)
                st.download_button(
                    label="Download Backtest Results as CSV",
                    data=csv,
                    file_name=f"{backtest_ticker}_backtest_results.csv",
                    mime="text/csv"
                )
else:
    st.info("Select stocks and fetch financial data first to enable backtesting.")

# Display Footer
st.markdown("---")
st.markdown("""
    <div style="text-align: center; padding: 20px; background-color: #1a1a1a; border-radius: 10px; margin-top: 30px;">
        <div style="display: flex; justify-content: center; align-items: center; margin-bottom: 15px;">
            <img src="https://predictram.com/images/logo.png" alt="Footer Logo" style="width: 90px; margin-right: 20px;">
            <div>
                <p style="font-size: 14px; margin: 5px 0;"><strong>App Code:</strong> Stock-Dividend-Prediction-Jan-2025</p>
                <p style="font-size: 14px; margin: 5px 0;">To get access to the stocks file, email us at <a href="mailto:support@pystatiq.com" style="color: #4CAF50;">support@pystatiq.com</a></p>
                <p style="font-size: 14px; margin: 5px 0;"><strong>Documentation:</strong> <a href="https://pystatiq-lab.gitbook.io/docs/python-apps/stock-dividend-predictions" target="_blank" style="color: #4CAF50;">Stock Dividend Prediction Guide</a></p>
            </div>
        </div>
        <p style="font-size: 12px; color: #9e9e9e; margin-top: 15px;">© 2025 PyStatIQ | Financial Analytics Platform</p>
    </div>
""", unsafe_allow_html=True)
